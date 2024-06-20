#  Copyright 2022-2024 EPAM Systems, Inc. (https://www.epam.com/)
#  #
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#  #
#     http://www.apache.org/licenses/LICENSE-2.0
#  #
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#

import datetime
import json
import os
import re
import traceback

from sls.app.storage_permissions_manager import StoragePermissionsManager
from sls.app.synchronizer.storage_synchronizer_interface import StorageLifecycleSynchronizer
from sls.app.model.sync_event_model import StorageLifecycleRuleActionItems
from sls.cloud import cloud_utils
from sls.pipelineapi.model.archive_rule_model import LifecycleRuleParser, StorageLifecycleRuleTransition, StorageLifecycleRuleProlongation
from sls.util import path_utils, date_utils

CRITERION_MATCHING_FILES = "MATCHING_FILES"
CRITERION_DEFAULT = "DEFAULT"

ACTIONS_MODE_FILES = "FILES"
ACTIONS_MODE_FOLDER = "FOLDER"

METHOD_ONE_BY_ONE = "ONE_BY_ONE"
METHOD_EARLIEST_FILE = "EARLIEST_FILE"
METHOD_LATEST_FILE = "LATEST_FILE"

EXECUTION_NOTIFICATION_SENT_STATUS = "NOTIFICATION_SENT"
EXECUTION_RUNNING_STATUS = "RUNNING"
EXECUTION_SUCCESS_STATUS = "SUCCESS"
EXECUTION_FAILED_STATUS = "FAILED"

ROLE_ADMIN_ID = 1


class StorageLifecycleArchivingSynchronizer(StorageLifecycleSynchronizer):

    def _sync_storage(self, storage):
        if storage.shared:
            self.logger.log("Storage {} marked as shared, skipping".format(storage.path))

        if self.config.rules_spec_file:
            if not os.path.isfile(self.config.rules_spec_file):
                self.logger.log("Custom rules file is specified but does not exist {}, skipping".format(self.config.rules_spec_file))
                return
            try:
                with open(self.config.rules_spec_file) as spec_file:
                    rules_json = json.load(spec_file)
                    rules_parser = LifecycleRuleParser()
                    rules = [rules_parser.parse_rule(rule, None, self.config.dry_run) for rule in rules_json if rules_json]

            except:
                self.logger.log("Failed reading custom rules file {}, skipping".format(self.config.rules_spec_file))
                return
        else:
            rules = self.pipeline_api_client.load_lifecycle_rules_for_storage(storage.id)

        # No rules for storage exist - just skip it
        if not rules:
            self.logger.log("No rules for storage {} is defined, skipping".format(storage.path))
            return

        if not self.cloud_bridge.is_support(storage):
            self.logger.log(
                "Lifecycle rules feature is not implemented for storage with type {}.".format(storage.storage_type)
            )
            return

        file_listing_cache = {}
        if not self.config.dry_run:
            self.cloud_bridge.prepare_bucket_if_needed(storage)
        for rule in rules:
            self.logger.log("Storage: {}. Rule: {}. [Starting]".format(storage.id, rule.rule_id))
            try:
                if self._rule_is_not_valid(rule, self.config.dry_run):
                    continue

                path_prefix = path_utils.determinate_prefix_from_glob(rule.path_glob)
                existing_listing_prefix = next(
                    filter(lambda k: path_prefix.startswith(k), file_listing_cache.keys()), None
                )
                if existing_listing_prefix:
                    files = [f for f in file_listing_cache[existing_listing_prefix] if f.path.startswith(path_prefix)]
                else:
                    files = self.cloud_bridge.list_objects_by_prefix(storage, path_prefix)
                    file_listing_cache[path_prefix] = files

                if self.config.dry_run:
                    running_executions = set()
                else:
                    running_executions = set(self.pipeline_api_client.load_lifecycle_rule_executions(
                        rule.datastorage_id, rule.rule_id, status=EXECUTION_RUNNING_STATUS))
                subject_folders = set(self._identify_subject_folders(files, rule.path_glob))
                subject_folders.update(e.path for e in running_executions)
                self.logger.log(
                    "Storage: {}. Rule: {}. Subject folders are: {}".format(storage.id, rule.rule_id, subject_folders))

                for folder in subject_folders:
                    self.logger.log(
                        "Storage: {}. Rule: {}. Path: '{}'. [Applying rule]".format(storage.id, rule.rule_id, folder))

                    # to match object keys with glob we need to combine it with folder path
                    # and get 'effective' glob like '{folder-path}/{glob}'
                    # so with glob = '*.txt' we can match files like:
                    # {folder-path}/{filename}.txt
                    effective_glob = os.path.join(folder, rule.object_glob) \
                        if rule.object_glob \
                        else os.path.join(folder, "*")

                    self.logger.log(
                        "Storage: {}. Rule: {}. Path: '{}'. Effective glob is '{}'".format(
                            storage.id, rule.rule_id, folder, effective_glob)
                    )

                    rule_subject_files = [
                        file for file in files
                        if file.path.startswith(folder)
                        and re.compile(path_utils.convert_glob_to_regexp(effective_glob)).match(file.path)
                    ] if effective_glob else files
                    self.logger.log(
                        "Storage: {}. Rule: {}. Path: '{}'. Found {} subject files, "
                        "than might be eligible to transition.".format(
                            storage.id, rule.rule_id, folder, len(rule_subject_files))
                    )

                    self._process_files(storage, folder, files, rule_subject_files, rule)
                    self.logger.log("Storage: {}. Rule: {}, Path: '{}'. [Rule applied]".format(storage.id, rule.rule_id, folder))
                self.logger.log("Storage: {}. Rule: {}. [Complete]".format(storage.id, rule.rule_id))
            except Exception as e:
                self.logger.log(
                    "Storage: {}. Rule: {}. Problems to apply the rule. "
                    "Cause: {}\n{}".format(storage.id, rule.rule_id, str(e), traceback.format_exc()))

    def _process_files(self, storage, folder, file_listing, rule_subject_files, rule):
        transition_method = rule.transition_method

        self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Transition method is {}. Building action items."
                        .format(rule.datastorage_id, rule.rule_id, folder, transition_method))

        if transition_method == METHOD_ONE_BY_ONE:
            resulted_action_items = self._build_action_items_for_files(folder, rule_subject_files, rule)
        else:
            resulted_action_items = self._build_action_items_for_folder(storage, folder, file_listing, rule_subject_files, rule)

        self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Done with building action items."
                        .format(rule.datastorage_id, rule.rule_id, folder))

        self._apply_action_items(storage, rule, resulted_action_items, dry_run=self.config.dry_run)

    def _build_action_items_for_files(self, folder, rule_subject_files, rule):
        result = StorageLifecycleRuleActionItems().with_folder(folder)\
            .with_mode(ACTIONS_MODE_FILES).with_rule_id(rule.rule_id)

        for file in rule_subject_files:
            effective_transitions = self._define_effective_transitions(
                self._fetch_prolongation_for_path_or_default(file.path, rule), rule.transitions)

            transition_class, transition_date = None, None
            for transition in effective_transitions:
                date_of_action = self._calculate_transition_date(file, transition)
                today = self.config.get_estimate_for_date()

                storage_class_file_filter = cloud_utils.get_storage_class_specific_file_filter(transition.storage_class)
                if date_utils.is_date_after_that(date_of_action, today):
                    should_be_transferred = \
                        file.storage_class != transition.storage_class and \
                        (transition_date is None or date_utils.is_date_after_that(transition_date, date_of_action)) and \
                        storage_class_file_filter(file)
                    if not should_be_transferred:
                        continue
                    transition_class = transition.storage_class
                    transition_date = date_of_action
            if transition_class:
                result.with_transition(transition_class, file)
        return result

    def _build_action_items_for_folder(self, storage, path, files_listing, subject_files_listing, rule):
        result = StorageLifecycleRuleActionItems().with_folder(path).with_rule_id(rule.rule_id)

        if self.config.dry_run:
            rule_executions = []
        else:
            rule_executions = self.pipeline_api_client.load_lifecycle_rule_executions(
                rule.datastorage_id, rule.rule_id, path=path)

        storage_class_transition_map = self.cloud_bridge.get_storage_class_transition_map(storage, rule)

        files_by_dest_storage_class = {}
        for dest_storage_class, source_storage_classes in storage_class_transition_map.items():
            for file in subject_files_listing:
                if file.storage_class in source_storage_classes:
                    files_by_dest_storage_class.setdefault(dest_storage_class, []).append(file)

        criterion_files_by_dest_storage_class = {
            dest_storage_class: self._define_criterion_file(rule, files_listing, path, files)
            for dest_storage_class, files in files_by_dest_storage_class.items()
        }

        self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Found {} executions.".format(
            rule.datastorage_id, rule.rule_id, path, len(rule_executions)))

        effective_transitions = self._define_effective_transitions(
            self._fetch_prolongation_for_path_or_default(path, rule), rule.transitions)

        for execution in rule_executions:
            transition = next(filter(lambda t: t.storage_class == execution.storage_class, effective_transitions), None)
            updated_execution = self._check_rule_execution_progress(
                rule.datastorage_id, transition, files_by_dest_storage_class, execution
            )
            if updated_execution:
                execution.status = updated_execution.status
                execution.updated = updated_execution.updated
                result.with_execution(updated_execution)

        today = self.config.get_estimate_for_date()

        transitions_by_dates = []
        for transition in effective_transitions:
            criterion_file = criterion_files_by_dest_storage_class.get(transition.storage_class, None)
            if criterion_file:
                transitions_by_dates.append((transition, self._calculate_transition_date(criterion_file, transition)))
        transitions_by_dates = sorted(transitions_by_dates, key=lambda pair: pair[1])

        for transition, transition_date in transitions_by_dates:
            transition_class = transition.storage_class
            transition_execution = next(
                filter(lambda e: e.storage_class == transition.storage_class, rule_executions), None
            )

            if transition_execution and transition_execution.status == EXECUTION_RUNNING_STATUS:
                self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Transition: {}. "
                                "No need for action: execution is in RUNNING state.".format(
                                 rule.datastorage_id, rule.rule_id, path, transition_class))
                continue

            trn_subject_files = self._filter_files_to_transit(files_by_dest_storage_class, transition, transition_date)
            if not trn_subject_files:
                self.logger.log(
                    "Storage: {}. Rule: {}. Path: '{}'. Transition: {}. No files to transit.".format(
                        rule.datastorage_id, rule.rule_id, path, transition_class))
                continue

            # Check if notification is needed
            notification = rule.notification
            if not self.config.dry_run and self._notification_should_be_sent(notification, transition_execution, transition_date, today):
                self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Transition: {}. Notification will be sent.".format(
                    rule.datastorage_id, rule.rule_id, path, transition_class, len(trn_subject_files)))
                result.with_notification(path, transition_class, str(transition_date), notification.prolong_days)

            # Check if action is needed
            if date_utils.is_date_after_that(transition_date, today):
                self.logger.log(
                    "Storage: {}. Rule: {}. Path: '{}'. Transition: {}. All criteria are met. Will transit {} files."
                    .format(rule.datastorage_id, rule.rule_id, path, transition_class, len(trn_subject_files)))

                for file in trn_subject_files:
                    result.with_transition(transition_class, file)
            else:
                self.logger.log(
                    "Storage: {}. Rule: {}. Path: '{}'. No eligible transition found now. "
                    "Nearest transition: destination - {} date - {}"
                    .format(rule.datastorage_id, rule.rule_id, path, transition_class, transition_date))
                continue
        return result

    def _filter_files_to_transit(self, files_by_dest_storage_class, transition, transition_date):
        trn_subject_files = files_by_dest_storage_class.get(transition.storage_class, None)
        if transition.transition_date is not None:
            trn_subject_files = [
                trn_file for trn_file in trn_subject_files if trn_file.creation_date.date() < transition_date
            ]

        storage_class_file_filter = cloud_utils.get_storage_class_specific_file_filter(transition.storage_class)
        trn_subject_files = [trn_file for trn_file in trn_subject_files if storage_class_file_filter(trn_file)]

        return trn_subject_files

    def _apply_action_items(self, storage, rule, action_items, dry_run=False):
        if dry_run:
            self._apply_action_items_dry_run(storage, rule, action_items)
        else:
            self._apply_action_items_real(storage, rule, action_items)

    def _apply_action_items_dry_run(self, storage, rule, action_items):
        # Just a report
        import openpyxl
        if os.path.isfile(self.config.dry_run_report_path):
            wb = openpyxl.load_workbook(self.config.dry_run_report_path)
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        if storage.name in wb.sheetnames:
            storage_sheet = wb[storage.name]
        else:
            storage_sheet = wb.create_sheet(storage.name)
            storage_sheet.append(['Storage ID',
                                'Storage Name',
                                'Source Tier',
                                'Destination Tier',
                                'Folder',
                                'File Path',
                                'File Creation Date',
                                'File Size'])
        
        for dest_tier in action_items.destination_transitions_queues.keys():
            for object_item in action_items.destination_transitions_queues[dest_tier]:
                row = [ str(storage.id), storage.name ]
                row.extend([object_item.storage_class,
                            dest_tier,
                            action_items.folder,
                            object_item.path,
                            str(object_item.creation_date),
                            str(object_item.size)])
                storage_sheet.append(row)
            
        wb.save(self.config.dry_run_report_path)

    def _apply_action_items_real(self, storage, rule, action_items):
        self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Performing action items."
                        .format(rule.datastorage_id, rule.rule_id, action_items.folder))

        for notification_properties in action_items.notification_queue:
            self._send_notification(storage, rule, notification_properties)

        for execution in action_items.executions:
            self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Transition: {}. Execution finished with status: {}"
                            .format(rule.datastorage_id, rule.rule_id, action_items.folder,
                                    execution.storage_class, execution.status))

        for storage_class, subject_files in action_items.destination_transitions_queues.items():
            if subject_files:
                self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Transition: {}. "
                                "Starting to tagging files for transition, number of files: {}.".format(
                    rule.datastorage_id, rule.rule_id, action_items.folder, storage_class, len(subject_files)))

                is_successful = self.cloud_bridge.tag_files_to_transit(
                    storage, subject_files, storage_class, self._get_transition_id(action_items, rule))

                if is_successful:
                    if action_items.mode == ACTIONS_MODE_FOLDER:
                        self._create_or_update_execution(storage.id, rule, storage_class,
                                                         action_items.folder, EXECUTION_RUNNING_STATUS)

        self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Complete action items."
                        .format(rule.datastorage_id, rule.rule_id, action_items.folder))

    def _create_or_update_execution(self, storage_id, rule, storage_class, path, status):
        executions = self.pipeline_api_client.load_lifecycle_rule_executions(storage_id, rule.rule_id, path=path)
        execution = next(filter(lambda e: e.storage_class == storage_class, executions), None)

        if execution:
            self.pipeline_api_client.update_status_lifecycle_rule_execution(
                storage_id, execution.execution_id, status)
        else:
            self.pipeline_api_client.create_lifecycle_rule_execution(
                storage_id, rule.rule_id,
                {
                    "status": status,
                    "path": path,
                    "storageClass": storage_class
                }
            )

    def _check_rule_execution_progress(self, storage_id, transition, subject_files_by_dest_storage_class, execution):
        if not execution.status:
            raise RuntimeError("Storage: {}. Rule: {}. Path: '{}'. Transition: {}. Malformed rule execution found."
                               " Status not found!".format(storage_id, execution.rule_id,
                                                           execution.path, execution.storage_class))

        self.logger.log(
            "Storage: {}. Rule: {}. Path: '{}'. Transition: {}. Checking existing execution, status: {}.".format(
                storage_id, execution.rule_id, execution.path, execution.storage_class, execution.status)
        )

        if execution.status != EXECUTION_RUNNING_STATUS:
            return None

        # Check that all files were moved to destination storage class and if not and 2 days are passed since
        # process was started - fail this execution
        subject_files = subject_files_by_dest_storage_class[execution.storage_class] \
            if execution.storage_class in subject_files_by_dest_storage_class \
            else []

        # So if we found a file in source location,
        # and creation date is before execution start date
        # (if transition was confgured with date - file should be created before that date) - then action is failed
        storage_class_file_filter = cloud_utils.get_storage_class_specific_file_filter(transition.storage_class)
        file_in_wrong_location = next(
            filter(lambda file: file.creation_date < execution.updated and
                                (transition.transition_date is None or file.creation_date.date() < transition.transition_date) and
                                storage_class_file_filter(file),
                   subject_files
            ), None
        )

        all_files_moved = file_in_wrong_location is None
        max_running_days = self.config.execution_max_running_days
        if all_files_moved:
            self.logger.log(
                "Storage: {}. Rule: {}. Path: '{}'. Transition: {}. "
                "All files moved to destination location, completing the action.".format(
                    storage_id, execution.rule_id, execution.path, execution.storage_class)
            )
            return self.pipeline_api_client.update_status_lifecycle_rule_execution(
                storage_id, execution.execution_id, EXECUTION_SUCCESS_STATUS)
        elif execution.updated + datetime.timedelta(days=max_running_days) < datetime.datetime.now(datetime.timezone.utc):
            self.logger.log(
                "Storage: {}. Rule: {}. Path: '{}'. Transition: {}. "
                "{} days are left but files in a wrong destination: {}. Failing this execution.".format(
                    storage_id, execution.rule_id, execution.path, execution.storage_class,
                    str(max_running_days), file_in_wrong_location.storage_class)
            )
            return self.pipeline_api_client.update_status_lifecycle_rule_execution(
                storage_id, execution.execution_id, EXECUTION_FAILED_STATUS)
        else:
            self.logger.log(
                "Storage: {}. Rule: {}. Path: '{}'. Transition: {}. Still RUNNING."
                .format(storage_id, execution.rule_id, execution.path, execution.storage_class)
            )
            return None

    def _send_notification(self, storage, rule, notification_properties):
        def _prepare_massage():
            cc_users = []
            for recipient in rule.notification.recipients:
                if recipient["principal"]:
                    cc_users.append(recipient["name"])
                else:
                    loaded_role = self.pipeline_api_client.load_role_by_name(recipient["name"])
                    if loaded_role and "users" in loaded_role:
                        cc_users.extend([user["userName"] for user in loaded_role["users"]])
            if rule.notification.notify_users:
                storage_users = StoragePermissionsManager(self.pipeline_api_client, storage.id).get_users()
                cc_users.extend([user_name for user_name in storage_users if not cc_users.__contains__(user_name)])

            _to_user = next(iter(cc_users), None)
            is_date_expired = self._is_action_date_expired(notification_properties)
            date_of_action = date_utils.current_date_string() if is_date_expired \
                else notification_properties["date_of_action"]
            notification_parameters = {
                "storageName": storage.name,
                "storageId": storage.id,
                "ruleId": rule.rule_id,
                "path": notification_properties["path"],
                "storageClass": notification_properties["storage_class"],
                "dateOfAction": date_of_action,
                "prolongDays": notification_properties["prolong_days"],
                "isDateNotExpired": not is_date_expired,
                "notificationType": "DATASTORAGE_LIFECYCLE_ACTION",
                "notificationResources": [{
                    "entityId": storage.id,
                    "entityClass": "STORAGE",
                    "storagePath": notification_properties["path"],
                    "storageRuleId": rule.rule_id,
                }],
            }

            return rule.notification.subject, rule.notification.body, _to_user, cc_users, notification_parameters

        path = notification_properties["path"]
        storage_class = notification_properties["storage_class"]
        if not rule.notification.enabled:
            self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Transition: {}. Notification disabled."
                            "Notification disabled.".format(rule.datastorage_id, rule.rule_id, path, storage_class))
            return
        else:
            self.logger.log("Storage: {}. Rule: {}. Path: '{}'. Transition: {}. Sending notification."
                            .format(rule.datastorage_id, rule.rule_id, path, storage_class))

        self._create_or_update_execution(
            rule.datastorage_id, rule, storage_class, path, EXECUTION_NOTIFICATION_SENT_STATUS)
        subject, body, to_user, copy_users, parameters = _prepare_massage()
        if subject and body and to_user:
            result = self.pipeline_api_client.send_notification(subject, body, to_user, copy_users, parameters)
            if not result:
                self._create_or_update_execution(
                    rule.datastorage_id, rule, storage_class, path, EXECUTION_FAILED_STATUS)
                raise RuntimeError("Problem to send a notification for: {}".format(str(notification_properties)))
        else:
            self.logger.log("Storage: {}. Rule: {}. Path: '{}'. "
                            "Will not send notification because parameters are not present, "
                            "subject: {} body: {} to_user: {}"
                            .format(storage.id, rule.rule_id, path, subject, body, to_user))

    @staticmethod
    def _notification_should_be_sent(notification, execution, date_of_action, today):
        if not notification.enabled:
            return False

        date_to_check = today + datetime.timedelta(days=notification.notify_before_days)
        if date_utils.is_date_before_that(date_of_action, date_to_check):
            return False

        if execution:
            was_updated_before = date_utils.is_date_after_that(
                date=today - datetime.timedelta(days=notification.notify_before_days),
                to_check=execution.updated.date()
            )
            if execution.status == EXECUTION_RUNNING_STATUS or execution.status == EXECUTION_FAILED_STATUS:
                return False
            if execution.status == EXECUTION_NOTIFICATION_SENT_STATUS and was_updated_before:
                return False
        return True

    @staticmethod
    def _rule_is_not_valid(rule, dry_run=False):
        if not dry_run:
            if not rule.notification:
                raise RuntimeError("Rule: {}. No notification defined!".format(rule.rule_id))
            if rule.notification.enabled:
                if not rule.notification.body or not rule.notification.subject:
                    raise RuntimeError("Rule: {}. Enabled notification should have subject and body!".format(rule.rule_id))
        if not rule.transitions:
            raise RuntimeError("Rule: {}. Transitions cannot be None or empty!".format(rule.rule_id))
        if not rule.transition_criterion:
            raise RuntimeError("Rule: {}. Transition criterion cannot be None!".format(rule.rule_id))
        if rule.transition_criterion.type != CRITERION_DEFAULT and not rule.transition_criterion.value:
            raise RuntimeError("Rule: {}. Transition criterion with not DEFAULT type should have value!"
                               .format(rule.rule_id))
        method_possible_values = [METHOD_LATEST_FILE, METHOD_ONE_BY_ONE, METHOD_EARLIEST_FILE]
        if not rule.transition_method or rule.transition_method not in method_possible_values:
            raise RuntimeError("Rule: {}. Transition method should have one of possible values: {}!"
                               .format(rule.rule_id, method_possible_values))

    @staticmethod
    def _calculate_transition_date(criterion_file, transition):
        if transition.transition_date:
            date_of_action = transition.transition_date
        elif transition.transition_after_days is not None:
            date_of_action = \
                (criterion_file.creation_date + datetime.timedelta(days=transition.transition_after_days)).date()
        else:
            raise RuntimeError("Malformed transition: date or days should be present.")
        return date_of_action

    @staticmethod
    def _define_criterion_file(rule, file_listing, folder, subject_files):
        transition_method = rule.transition_method
        criterion_files = subject_files
        if rule.transition_criterion.type == CRITERION_MATCHING_FILES:
            transition_criterion_files_glob = os.path.join(folder, rule.transition_criterion.value)
            criterion_files = [
                file for file in file_listing
                if file.path.startswith(folder) and
                   re.compile(path_utils.convert_glob_to_regexp(transition_criterion_files_glob)).match(file.path)
            ]
        # Sort by date (reverse or not depends on transition method) and get first element or None
        return next(
            iter(
                sorted(
                    criterion_files, key=lambda e: e.creation_date, reverse=(transition_method == METHOD_LATEST_FILE)
                )
            ), None
        )

    @staticmethod
    def _define_effective_transitions(prolongation, transitions):
        def _define_transition(transition):
            resulted_transition = StorageLifecycleRuleTransition(transition.storage_class)
            if transition.transition_after_days is not None:
                resulted_transition.transition_after_days = transition.transition_after_days + prolongation.days
            if transition.transition_date:
                resulted_transition.transition_date = \
                    transition.transition_date + datetime.timedelta(days=prolongation.days)
            return resulted_transition

        return [_define_transition(transition) for transition in transitions]

    @staticmethod
    def _fetch_prolongation_for_path_or_default(path, rule):
        file_prolongation = StorageLifecycleRuleProlongation(prolongation_id=-1, prolonged_date=None, days=0, path=path)
        if rule.prolongations:
            file_prolongation = next(
                filter(lambda prolongation: prolongation.path == path, rule.prolongations),
                file_prolongation
            )
        return file_prolongation

    @staticmethod
    def _get_transition_id(action_items, rule):
        return "archive_rule_{}_{}".format(str(rule.rule_id), action_items.folder).replace(" ", "_").replace(":", "_").replace("/", ".")

    @staticmethod
    def _identify_subject_folders(files, glob_str):
        if not glob_str or glob_str == "/":
            return ["/"]

        unique_folder_paths = set()
        for file in files:
            parent_dir, filename = os.path.split(file.path)
            if parent_dir:
                unique_folder_paths.add(parent_dir)
        directories = path_utils.generate_all_possible_dir_paths(unique_folder_paths)
        resulted_regexp = path_utils.convert_glob_to_regexp(glob_str)
        pattern = re.compile(resulted_regexp)
        return [p for p in directories if pattern.match(p)]

    @staticmethod
    def _is_action_date_expired(notification_properties):
        date_of_action = notification_properties["date_of_action"]
        if not date_of_action:
            return False
        return date_utils.is_date_before_now(date_utils.parse_date(date_of_action))
